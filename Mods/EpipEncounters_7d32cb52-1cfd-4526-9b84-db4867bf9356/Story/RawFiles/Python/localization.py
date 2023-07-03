import re, os, xml, json
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils import *
from openpyxl.styles import *
from openpyxl.worksheet import worksheet

TEXTLIB_TRANSLATION_FILE_FORMAT_VERSION = 0

OUTDATED_TRANSLATION_FILL = 'FFF03A3A'

# -----------------
# CLASSES
# -----------------

class TranslatedString:
    def __init__(self, handle, obj):
        self.text = obj["ReferenceText"]
        self.description = obj["ContextDescription"] if "ContextDescription" in obj else ""
        self.translation = obj["TranslatedText"] if "TranslatedText" in obj else ""
        self.key = obj["ReferenceKey"] if "ReferenceKey" in obj else ""
        self.handle = handle
        self.featureID = obj["FeatureID"] if "FeatureID" in obj else ""

        self.TranslationNotes = ""
        self.TranslationAuthors = ""
        self.outdated_translation = False

    def has_translation(self)->bool:
        return self.translation != ""

class SheetColumn:
    def __init__(self, name, width, tsk_field:str, locked:bool=True, wrap_text:bool=True, font=None, json_key:str=None):
        self.name = name
        self.width = width
        self.locked = locked
        self.wrap_text = wrap_text
        self.font = font
        self.json_key = json_key
        self.tsk_field = tsk_field

def addSheetRow(column_defs, i, sheet, row, content:str):
    column_def = column_defs[i]
    letter = get_column_letter(i + 1)

    cell = sheet[letter + str(row)]

    sheet.row_dimensions[row].height = Sheet.ROW_HEIGHT
    
    # Set font
    cell.font = column_def.font

    cell.value = content

    # Set alignment
    if column_def.wrap_text:
        cell.alignment = Alignment(wrap_text=True)

    # Unlock cell for editing
    if not column_def.locked:
        cell.protection = Protection(locked=False)

class Sheet:
    ROW_HEIGHT = 30

    def __init__(self, sheet, name):
        self.name = name
        self.sheet = sheet
        self.column_defs = []
        self._initialized = False
        self.filled_rows = 0

        self.sheet.protection.sheet = True

    def addColumnDefinition(self, column:SheetColumn):
        self.column_defs.append(column)

        # Set width
        self.sheet.column_dimensions[get_column_letter(len(self.column_defs))].width = column.width

        # Protection is set when adding cell content instead.

    def append(self, values):
        self.sheet.append(values)

    def addRow(self, values:list):
        # Add the header row
        if not self._initialized:
            self.sheet.append([column.name for column in self.column_defs])

            self._initialized = True

        self.filled_rows = self.filled_rows + 1

        if len(values) != len(self.column_defs):
            raise ValueError("Iterable length mismatch with column defs amount")

        self.sheet.append(values)

        for i in range(len(values)):
            addSheetRow(self.column_defs, i, self.sheet, self.filled_rows + 1, values[i])


class Workbook:
    def __init__(self):
        self.book = openpyxl.Workbook()
        self.sheets = []
        self._initalized = False

    def addSheet(self, name):
        sheet = Sheet(self.book.create_sheet(name), name)
        self.sheets.append(sheet)

        if not self._initalized:
            self.book.remove(self.book["Sheet"]) # TODO maybe don't?
            self._initalized = True

        return sheet

    def save(self, file_name):
        self.book.save(file_name)

class Module:
    COLUMN_DEFINITIONS = [
        SheetColumn("Handle", 10, "handle", wrap_text=False, font=Font(color="999999")),
        SheetColumn("Script", 20, "featureID", wrap_text=False),
        SheetColumn("Original Text", 65, "text", json_key="ReferenceText"),
        SheetColumn("Contextual Info", 40, "description"),
        SheetColumn("Translation", 65, "translation", locked=False, json_key="TranslatedText"),
        SheetColumn("Translation Notes", 30, "TranslationNotes", locked=False),
        SheetColumn("Translation Author", 30, "TranslationAuthors", locked=False),
    ]

    def __init__(self, name):
        self.name = name
        self.translated_strings:dict[str, TranslatedString] = {}

    def addTSK(self, tsk):
        self.translated_strings[tsk.handle] = tsk

    def get_tsk(self, handle)->TranslatedString:
        return self.translated_strings[handle]

    def update_tsk_row(self, handle, sheet, row_index):
        tskData = self.get_tsk(handle)
        translation_cell = sheet.cell(row=row_index, column=5)

        # Setup row properties
        sheet.row_dimensions[row_index].height = Sheet.ROW_HEIGHT

        # Setup properties of each cell within the row
        for i in range(len(Module.COLUMN_DEFINITIONS)):
            column_def = Module.COLUMN_DEFINITIONS[i]
            cell = sheet.cell(row=row_index, column=i+1)

            cell.font = column_def.font
            if column_def.wrap_text:
                cell.alignment = Alignment(wrap_text=True)
            cell.protection = Protection(locked=column_def.locked)

        # Recolor cells as a warning when the old original text doesn't match new one
        if tskData.outdated_translation and tskData.has_translation():
            print("Outdated translation for", tskData.handle, tskData.text)

            translation_cell.fill = PatternFill(start_color=OUTDATED_TRANSLATION_FILL, end_color=OUTDATED_TRANSLATION_FILL, fill_type='solid')

        # Update all columns
        for i in range(len(Module.COLUMN_DEFINITIONS)):
            column = Module.COLUMN_DEFINITIONS[i]
            cell = sheet.cell(row=row_index, column=i+1)

            cell.value = getattr(tskData, column.tsk_field)

    def export(self, book:Workbook, oldBook:openpyxl.Workbook=None, language_sheet:str=None, oldBookPath:str=None):
        if oldBook == None:
            sheet = book.addSheet(self.name)

            for column_def in Module.COLUMN_DEFINITIONS:
                sheet.addColumnDefinition(column_def)

            for handle,tsk in self.translated_strings.items():
                sheet.addRow([
                    handle,
                    tsk.featureID,
                    tsk.text,
                    tsk.description,
                    tsk.translation, # Translation
                    "", # Translation author - these are not kept within the json, cannot import
                    "", # Translation notes - same case as author
                ])
        else:
            print("Parsing sheet", language_sheet)

            oldSheet:worksheet = oldBook.get_sheet_by_name(language_sheet)

            existingTSKs = set()
            for rowIndex in range(2, oldSheet.max_row + 1): # Ignore first row (header)
                handleCell = oldSheet.cell(row=rowIndex, column=1)
                handle = handleCell.value

                if handle != None:
                    tskExists = handle in self.translated_strings

                    existingTSKs.add(handle)

                    # Retrieve extra TSK data from sheet
                    if tskExists:
                        tsk = self.get_tsk(handle)
                        tsk.TranslationNotes = oldSheet.cell(row=rowIndex, column=6).value
                        tsk.TranslationAuthors = oldSheet.cell(row=rowIndex, column=7).value
                        tsk.outdated_translation = oldSheet.cell(row=rowIndex, column=3).value != tsk.text
                    else:
                        print("Found removed TSK", handle, oldSheet.cell(row=rowIndex, column=3).value)

            # Sort and update all rows, including new ones
            sorted_tsks:list[TranslatedString] = list(self.translated_strings.values())
            sorted_tsks.sort(key=lambda x: x.featureID + "_" + x.text)

            for i in range(0, len(sorted_tsks)):
                tsk = sorted_tsks[i]
                self.update_tsk_row(tsk.handle, oldSheet, i + 2)

            oldBook.save(oldBookPath)

def createSpreadsheet(file_name, existing_sheet_file_name=None, language_sheet:str=None):
    localization = json.load(open(file_name, "r"))
    module = Module(localization["ModTable"])

    for handle,data in localization["TranslatedStrings"].items():
        tsk = TranslatedString(handle, data)

        module.addTSK(tsk)

    oldBook = None
    if existing_sheet_file_name != None:
        oldBook = openpyxl.open(existing_sheet_file_name)
    book = Workbook()

    print("Saving workbook")
    module.export(book, oldBook, language_sheet, existing_sheet_file_name)

def createTranslationJSON(file_name:str, modTable:str, sheet_name:str, output_filename:str):
    book = openpyxl.load_workbook(file_name)
    sheet = book[sheet_name]

    output = {
        "FileFormatVersion": TEXTLIB_TRANSLATION_FILE_FORMAT_VERSION,
        "ModTable": modTable,
        "TranslatedStrings": {},
    }
    tsks = output["TranslatedStrings"]

    for row in range(2, sheet.max_row + 1):
        handle = sheet["A" + str(row)].value
        tsks[handle] = {}
        tsk = tsks[handle]

        for column_index in range(len(Module.COLUMN_DEFINITIONS)):
            column_def = Module.COLUMN_DEFINITIONS[column_index]

            if column_def.json_key:
                cell = sheet[get_column_letter(column_index + 1) + str(row)]
                value = cell.value

                tsk[column_def.json_key] = value

        # erase TSKs with no translation
        handles_to_delete = []
        for handle,data in tsks.items():
            if "TranslatedText" not in data or data["TranslatedText"] == None:
                handles_to_delete.append(handle)

        for handle in handles_to_delete:
            del tsks[handle]

    with open(output_filename, "w") as f:
        f.write(json.dumps(output, indent=2, ensure_ascii=False))

        print("Converted sheet to json")